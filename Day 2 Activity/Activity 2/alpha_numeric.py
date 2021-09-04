import re
from openpyxl import load_workbook   #install openpyxl first : pip install openpyxl
file_name = "Alphanumeric.xlsx"
wb = load_workbook(file_name, read_only=True) 
ws = wb[wb.sheetnames[0]]

#Row and Column size of excel sheet

row_count = ws.max_row
column_count = ws.max_column

#regex

reg='[A-Za-z]+'

#print values

for i in range(2, row_count+1):
    match = re.findall(reg, ws["A{0}".format(i)].value )
    print(match)